"""
report_generator.py — STEP 6
분석 결과 → 엑셀 대시보드 + PDF 리포트 자동 생성 (한글 폰트 지원)
"""

import os
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

DESKTOP_PATH    = os.path.join(os.path.expanduser("~"), "Desktop")
EXCEL_FILE_NAME = os.path.join(DESKTOP_PATH, "instagram_analysis_result.xlsx")

# ── 한글 폰트 등록 ────────────────────────────────────────────
def _register_korean_font() -> tuple[str, str]:
    """Regular + Bold 폰트 등록. (regular_name, bold_name) 반환"""
    if sys.platform == "win32":
        candidates = [
            (r"C:\Windows\Fonts\malgun.ttf",   r"C:\Windows\Fonts\malgunbd.ttf"),
            (r"C:\Windows\Fonts\NanumGothic.ttf", r"C:\Windows\Fonts\NanumGothicBold.ttf"),
        ]
    elif sys.platform == "darwin":
        candidates = [
            ("/Library/Fonts/NanumGothic.ttf", "/Library/Fonts/NanumGothicBold.ttf"),
        ]
    else:
        candidates = [
            ("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
             "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc"),
            ("/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
             "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"),
        ]

    for reg_path, bold_path in candidates:
        if os.path.exists(reg_path):
            try:
                pdfmetrics.registerFont(TTFont("KR", reg_path))
                if os.path.exists(bold_path):
                    pdfmetrics.registerFont(TTFont("KR-Bold", bold_path))
                    return "KR", "KR-Bold"
                return "KR", "KR"
            except Exception as e:
                print(f"     폰트 등록 실패: {e}")
                continue

    return "Helvetica", "Helvetica-Bold"

KR, KR_BOLD = _register_korean_font()

# ── 색상 ─────────────────────────────────────────────────────
CLR = {
    "navy":       "1A1A2E",
    "red":        "E94560",
    "gold":       "F5A623",
    "green":      "27AE60",
    "yellow":     "F39C12",
    "danger":     "E74C3C",
    "light":      "F8F9FA",
    "white":      "FFFFFF",
    "gray":       "95A5A6",
    "light_gray": "ECF0F1",
}
GRADE_CLR = {"S": "8E44AD", "A": "27AE60", "B": "F39C12", "C": "E67E22", "D": "E74C3C"}
RISK_CLR  = {"LOW": "27AE60", "MEDIUM": "F39C12", "HIGH": "E74C3C"}


# ══════════════════════════════════════════════════════════════
# 📊 엑셀 대시보드
# ══════════════════════════════════════════════════════════════
def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _c(ws, row, col, value, bold=False, size=10, color="000000",
       bg=None, align="left", border=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if bg:  c.fill = PatternFill("solid", fgColor=bg)
    if border: c.border = _thin()
    if num_fmt: c.number_format = num_fmt
    return c

def save_excel_dashboard(username, score_result, account_data, sample_size):
    wb = load_workbook(EXCEL_FILE_NAME) if os.path.exists(EXCEL_FILE_NAME) else Workbook()

    # ── 시트1: 누적 로그 ──────────────────────────────────────
    LOG = "분석 로그"
    if LOG not in wb.sheetnames:
        ws = wb.create_sheet(LOG)
        headers = [
            "분석일시","계정명","최종점수","등급","팔로워티어","팔로워수",
            "카테고리","평균좋아요","원시ER(%)","진성댓글(%)","봇댓글(%)",
            "가짜팔로워(%)","리스크","상태","피드단가","스토리단가","릴스단가",
            "참여도","팔로워품질","댓글품질","일관성",
        ]
        for i, h in enumerate(headers, 1):
            _c(ws, 1, i, h, bold=True, color="FFFFFF",
               bg=CLR["navy"], align="center", border=True)
            ws.column_dimensions[get_column_letter(i)].width = max(len(h)+4, 12)
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
    else:
        ws = wb[LOG]

    cq  = account_data.get("comment_quality", {})
    cat = score_result.get("category", {})
    adp = score_result.get("ad_price", {})
    sb  = score_result.get("score_breakdown", {})
    r   = ws.max_row + 1
    bg  = CLR["light"] if r % 2 == 0 else CLR["white"]

    row_data = [
        datetime.now().strftime("%Y-%m-%d %H:%M"), username,
        score_result["final_score"], score_result["grade"], score_result["tier"],
        account_data.get("follower_count", 0),
        cat.get("primary", "일반"),
        round(account_data.get("avg_likes_per_post", 0), 1),
        score_result["raw_er"],
        round(cq.get("genuine_ratio", 0)*100, 1),
        round(cq.get("bot_ratio", 0)*100, 1),
        score_result["fake_ratio"], score_result["fake_follower_risk"],
        score_result["status"],
        adp.get("feed_fmt","-"), adp.get("story_fmt","-"), adp.get("reels_fmt","-"),
        sb.get("engagement",0), sb.get("follower_quality",0),
        sb.get("comment_quality",0), sb.get("consistency",0),
    ]
    for i, val in enumerate(row_data, 1):
        _c(ws, r, i, val, bg=bg, border=True, align="center")
    grade = score_result["grade"]
    ws.cell(r, 4).font = Font(name="Arial", bold=True,
                               color=GRADE_CLR.get(grade,"000000"), size=11)

    # ── 시트2: 최신 대시보드 ──────────────────────────────────
    DSH = "📊 대시보드"
    if DSH in wb.sheetnames: del wb[DSH]
    ds = wb.create_sheet(DSH, 0)
    ds.sheet_view.showGridLines = False
    for col in range(1, 12):
        ds.column_dimensions[get_column_letter(col)].width = 14
    ds.column_dimensions["A"].width = 2

    row = 2
    ds.merge_cells(f"B{row}:K{row}")
    _c(ds, row, 2, f"Instagram Influencer Report  |  @{username}",
       bold=True, size=15, color="FFFFFF", bg=CLR["navy"], align="center")
    ds.row_dimensions[row].height = 34
    row += 1

    ds.merge_cells(f"B{row}:K{row}")
    _c(ds, row, 2,
       f"분석일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  샘플: {sample_size}명  |  카테고리: {cat.get('primary','일반')}",
       size=9, color="AAAAAA", bg=CLR["navy"], align="center")
    ds.row_dimensions[row].height = 18
    row += 2

    # 카드 4개
    cards = [
        ("최종 점수",  f"{score_result['final_score']}점", GRADE_CLR.get(grade,"000000")),
        ("등급",       f"{grade} 등급",                    GRADE_CLR.get(grade,"000000")),
        ("팔로워 티어", score_result["tier"].upper(),       CLR["navy"]),
        ("리스크",     score_result["fake_follower_risk"],  RISK_CLR.get(score_result["fake_follower_risk"],"000000")),
    ]
    for idx, (label, val, clr) in enumerate(cards):
        c0 = 2 + idx*2
        ds.merge_cells(start_row=row,   start_column=c0, end_row=row,   end_column=c0+1)
        ds.merge_cells(start_row=row+1, start_column=c0, end_row=row+1, end_column=c0+1)
        _c(ds, row,   c0, label, size=8, color="888888", bg=CLR["light_gray"], align="center")
        _c(ds, row+1, c0, val,   bold=True, size=14, color=clr,
           bg=CLR["light_gray"], align="center")
        ds.row_dimensions[row].height   = 16
        ds.row_dimensions[row+1].height = 28
    row += 3

    # 세부 점수
    ds.merge_cells(f"B{row}:K{row}")
    _c(ds, row, 2, "  세부 점수", bold=True, size=11,
       color="FFFFFF", bg=CLR["red"], align="left")
    ds.row_dimensions[row].height = 22
    row += 1

    for label, key, desc in [
        ("참여도(ER)",   "engagement",       "ER 기반 티어 정규화"),
        ("팔로워 품질",  "follower_quality", "샘플 300명 신뢰도"),
        ("댓글 품질",    "comment_quality",  "진성 댓글 비율"),
        ("일관성",       "consistency",      "게시물 편차"),
    ]:
        v   = sb.get(key, 0)
        pct = round(v*100)
        bar = "█"*max(0,min(10,int(v*10))) + "░"*(10-max(0,min(10,int(v*10))))
        clr = CLR["green"] if pct>=70 else CLR["yellow"] if pct>=40 else CLR["danger"]
        ds.merge_cells(f"B{row}:C{row}")
        _c(ds, row, 2, label, bold=True, size=9, border=True)
        ds.merge_cells(f"D{row}:G{row}")
        _c(ds, row, 4, bar, size=9, color=clr, border=True)
        _c(ds, row, 8, f"{pct}점", bold=True, size=10, color=clr, align="center", border=True)
        ds.merge_cells(f"I{row}:K{row}")
        _c(ds, row, 9, desc, size=8, color="888888", border=True)
        ds.row_dimensions[row].height = 20
        row += 1
    row += 1

    # 광고 단가
    ds.merge_cells(f"B{row}:K{row}")
    _c(ds, row, 2, "  광고 단가 추정", bold=True, size=11,
       color="FFFFFF", bg=CLR["gold"], align="left")
    ds.row_dimensions[row].height = 22
    row += 1

    for icon, label, key in [("📸","피드","feed_fmt"), ("📖","스토리","story_fmt"), ("🎬","릴스","reels_fmt")]:
        ds.merge_cells(f"B{row}:D{row}")
        _c(ds, row, 2, f"{icon} {label}", bold=True, size=10, border=True)
        ds.merge_cells(f"E{row}:K{row}")
        _c(ds, row, 5, adp.get(key,"-"), bold=True, size=12,
           color=CLR["red"], border=True, align="center")
        ds.row_dimensions[row].height = 24
        row += 1

    wb.save(EXCEL_FILE_NAME)
    print(f"📊 엑셀 대시보드 저장: {EXCEL_FILE_NAME}")


# ══════════════════════════════════════════════════════════════
# 📄 PDF 리포트 (한글 폰트 적용)
# ══════════════════════════════════════════════════════════════
def save_pdf_report(username, score_result, account_data, sample_size) -> str:
    filename = os.path.join(
        DESKTOP_PATH,
        f"insta_report_{username}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    )
    doc = SimpleDocTemplate(filename, pagesize=A4,
                             topMargin=1.5*cm, bottomMargin=1.5*cm,
                             leftMargin=2*cm,  rightMargin=2*cm)

    navy  = colors.HexColor("#1A1A2E")
    red   = colors.HexColor("#E94560")
    gold  = colors.HexColor("#F5A623")
    gray  = colors.HexColor("#95A5A6")
    lgray = colors.HexColor("#ECF0F1")
    green = colors.HexColor("#27AE60")
    yel   = colors.HexColor("#F39C12")
    dred  = colors.HexColor("#E74C3C")

    def ps(name, font=KR, **kw):
        return ParagraphStyle(name, fontName=font, **kw)

    title_s = ps("T", font=KR_BOLD, fontSize=20, textColor=colors.white,
                  backColor=navy, alignment=TA_CENTER, leading=30,
                  spaceAfter=0, spaceBefore=0)
    sub_s   = ps("S", fontSize=9,  textColor=gray, alignment=TA_CENTER, spaceAfter=10)
    h2_s    = ps("H2", font=KR_BOLD, fontSize=12, textColor=colors.white,
                  backColor=red, leading=20, spaceBefore=12, spaceAfter=4)
    body_s  = ps("B", fontSize=9, textColor=colors.HexColor("#333333"), leading=14)
    foot_s  = ps("F", fontSize=7, textColor=gray, alignment=TA_CENTER)

    cq  = account_data.get("comment_quality", {})
    cat = score_result.get("category", {})
    adp = score_result.get("ad_price", {})
    sb  = score_result.get("score_breakdown", {})
    grade = score_result["grade"]
    gc    = colors.HexColor(f"#{GRADE_CLR.get(grade,'000000')}")

    story = []

    # 타이틀
    story.append(Paragraph("Instagram Influencer Report", title_s))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"@{username}  ·  {datetime.now().strftime('%Y년 %m월 %d일')}  ·  샘플 {sample_size}명",
        sub_s))
    story.append(HRFlowable(width="100%", thickness=1, color=red))
    story.append(Spacer(1, 10))

    # 핵심 카드
    def para(text, font=KR, size=10, color=colors.black, align=TA_CENTER):
        s = ps("tmp", font=font, fontSize=size, textColor=color, alignment=align, leading=size+4)
        return Paragraph(text, s)

    card_data = [[
        para(f"{score_result['final_score']}점\n최종 점수", KR_BOLD, 18, gc),
        para(f"{grade} 등급\n등급",                         KR_BOLD, 18, gc),
        para(f"{score_result['tier'].upper()}\n팔로워 티어", KR, 14, navy),
        para(f"{cat.get('primary','일반')}\n카테고리",       KR, 14, red),
    ]]
    ct = Table(card_data, colWidths=[4*cm]*4)
    ct.setStyle(TableStyle([
        ("ALIGN",       (0,0),(-1,-1),"CENTER"),
        ("VALIGN",      (0,0),(-1,-1),"MIDDLE"),
        ("ROWHEIGHT",   (0,0),(-1,-1), 55),
        ("BOX",         (0,0),(-1,-1), 0.5, colors.HexColor("#DDDDDD")),
        ("INNERGRID",   (0,0),(-1,-1), 0.5, colors.HexColor("#DDDDDD")),
        ("BACKGROUND",  (0,0),(-1,-1), lgray),
        ("TOPPADDING",  (0,0),(-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
    ]))
    story.append(ct)
    story.append(Spacer(1, 14))

    # 세부 점수
    story.append(Paragraph("세부 점수 분석", h2_s))
    story.append(Spacer(1, 4))

    def bar(v):
        f = max(0, min(10, int(v*10)))
        return "■"*f + "□"*(10-f)

    score_rows = [
        [para("지표",KR_BOLD,9,colors.white), para("점수",KR_BOLD,9,colors.white),
         para("바 차트",KR_BOLD,9,colors.white), para("설명",KR_BOLD,9,colors.white)],
    ]
    for label, key, desc in [
        ("참여도 (ER)",      "engagement",       "ER 기반 티어 정규화"),
        ("팔로워 품질",      "follower_quality", "샘플 300명 신뢰도"),
        ("댓글 품질",        "comment_quality",  "진성 댓글 비율"),
        ("인게이지먼트 일관성","consistency",     "게시물 간 편차"),
    ]:
        v   = sb.get(key, 0)
        pct = round(v*100)
        clr = green if pct>=70 else yel if pct>=40 else dred
        score_rows.append([
            para(label, KR, 9),
            para(f"{pct}점", KR_BOLD, 9, clr),
            para(bar(v), KR, 9, clr),
            para(desc, KR, 8, gray),
        ])

    st = Table(score_rows, colWidths=[4.5*cm, 2*cm, 4.5*cm, 5*cm])
    st.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), navy),
        ("FONTSIZE",      (0,0),(-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, lgray]),
        ("BOX",           (0,0),(-1,-1), 0.5, colors.HexColor("#DDDDDD")),
        ("INNERGRID",     (0,0),(-1,-1), 0.5, colors.HexColor("#EEEEEE")),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("ALIGN",         (1,0),(2,-1), "CENTER"),
    ]))
    story.append(st)
    story.append(Spacer(1, 14))

    # 가짜 팔로워
    story.append(Paragraph("가짜 팔로워 진단", h2_s))
    story.append(Spacer(1, 4))

    risk      = score_result["fake_follower_risk"]
    risk_bg   = {"LOW": colors.HexColor("#D5F5E3"),
                 "MEDIUM": colors.HexColor("#FDEBD0"),
                 "HIGH": colors.HexColor("#FADBD8")}.get(risk, lgray)

    fake_rows = [
        [para("항목",KR_BOLD,9,colors.white), para("수치",KR_BOLD,9,colors.white)],
        [para("가짜 팔로워 비율",KR,9),  para(f"{score_result['fake_ratio']}%",KR_BOLD,10)],
        [para("리스크 등급",KR,9),        para(risk, KR_BOLD, 10,
                                                colors.HexColor(f"#{RISK_CLR.get(risk,'000000')}"))],
        [para("진단 상태",KR,9),          para(score_result["status"],KR,9)],
        [para("진성 댓글 비율",KR,9),     para(f"{round(cq.get('genuine_ratio',0)*100,1)}%",KR,9)],
        [para("봇 댓글 비율",KR,9),       para(f"{round(cq.get('bot_ratio',0)*100,1)}%",KR,9)],
        [para("분석 샘플 수",KR,9),       para(f"{sample_size}명",KR,9)],
    ]
    ft = Table(fake_rows, colWidths=[8*cm, 8*cm])
    ft.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), navy),
        ("BACKGROUND",    (0,1),(-1,-1), risk_bg),
        ("FONTSIZE",      (0,0),(-1,-1), 9),
        ("BOX",           (0,0),(-1,-1), 0.5, colors.HexColor("#DDDDDD")),
        ("INNERGRID",     (0,0),(-1,-1), 0.5, colors.HexColor("#EEEEEE")),
        ("ALIGN",         (1,0),(1,-1), "CENTER"),
        ("TOPPADDING",    (0,0),(-1,-1), 6),
        ("BOTTOMPADDING", (0,0),(-1,-1), 6),
    ]))
    story.append(ft)
    story.append(Spacer(1, 14))

    # 광고 단가
    story.append(Paragraph("광고 단가 추정", ps("H3",font=KR_BOLD,fontSize=12,
                             textColor=colors.white, backColor=gold,
                             leading=20, spaceBefore=12, spaceAfter=4)))
    story.append(Spacer(1, 4))

    price_rows = [
        [para("광고 유형",KR_BOLD,9,colors.white),
         para("추정 단가",KR_BOLD,9,colors.white),
         para("비고",KR_BOLD,9,colors.white)],
        [para("피드 게시물",KR,10), para(adp.get("feed_fmt","-"),KR_BOLD,11,red),  para("1회 게시물",KR,8,gray)],
        [para("스토리",KR,10),      para(adp.get("story_fmt","-"),KR_BOLD,11,red), para("24시간 노출",KR,8,gray)],
        [para("릴스",KR,10),        para(adp.get("reels_fmt","-"),KR_BOLD,11,red), para("숏폼 영상",KR,8,gray)],
    ]
    pt = Table(price_rows, colWidths=[5*cm, 5.5*cm, 5.5*cm])
    pt.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), gold),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, lgray]),
        ("BOX",           (0,0),(-1,-1), 0.5, colors.HexColor("#DDDDDD")),
        ("INNERGRID",     (0,0),(-1,-1), 0.5, colors.HexColor("#EEEEEE")),
        ("ALIGN",         (1,0),(1,-1), "CENTER"),
        ("TOPPADDING",    (0,0),(-1,-1), 8),
        ("BOTTOMPADDING", (0,0),(-1,-1), 8),
    ]))
    story.append(pt)
    story.append(Spacer(1, 20))

    # 푸터
    story.append(HRFlowable(width="100%", thickness=0.5, color=gray))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"Generated by insta_engine v4.3  ·  {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        foot_s))

    doc.build(story)
    print(f"📄 PDF 저장: {filename}")
    return filename


def _bar_str(v: float) -> str:
    f = max(0, min(10, int(v*10)))
    return "■"*f + "□"*(10-f)


# ══════════════════════════════════════════════════════════════
# 통합 호출
# ══════════════════════════════════════════════════════════════
def generate_reports(username, score_result, account_data, sample_size):
    save_excel_dashboard(username, score_result, account_data, sample_size)
    return save_pdf_report(username, score_result, account_data, sample_size)